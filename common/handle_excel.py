"""
======================================
Author:包天宇
Time:2021-04-24 22:23
E-mail:uattb991492@icccuat.com
======================================
"""
import openpyxl
import os
from common.handle_path import DATA_DIR


class Excel:

    def __init__(self, filename, sheetname):
        """

        :param filename: excle的路径名
        :param sheetname: sheet页的名称
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        """
        用于读取excel中的用例
        :return:
        """
        workbook = openpyxl.load_workbook(self.filename)
        # 获取sheet页
        sh = workbook[self.sheetname]
        # rows:按行获取表单中所有的格子,每一行的格子放到一个元组中
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行之外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        # 返回读取出来的数据
        return cases

    def create_excel(self, row, col, value):
        """
        用于回显结果至excel
        :param row: 行数
        :param col: 列数
        :param value:  值
        :return:
        """
        excel = openpyxl.load_workbook(self.filename)
        sh = excel[self.sheetname]
        sh.cell(row=row, column=col, value=value)
        excel.save(self.filename)

